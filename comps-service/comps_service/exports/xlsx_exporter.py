from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from talk_to_your_stock_shared import CompsTable
from comps_service.exports.csv_exporter import COLUMNS


def to_xlsx(table: CompsTable) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trading Comps"
    sheet.append([column.replace("_", " ").title() for column in COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for row in table.rows:
        payload = row.model_dump(mode="json")
        sheet.append([payload.get(column) for column in COLUMNS])
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 28)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
